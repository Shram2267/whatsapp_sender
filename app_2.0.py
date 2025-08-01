import streamlit as st
import pandas as pd
import requests
import json
import os
import subprocess
import re
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from concurrent.futures import ThreadPoolExecutor, as_completed

TEMPLATE_FILE = "templates.json"

# ---------- Utilities ----------
@st.cache_resource
def get_templates_cached():
    return load_templates()

def load_templates():
    if os.path.exists(TEMPLATE_FILE):
        with open(TEMPLATE_FILE, "r") as f:
            return json.load(f)
    return []

def save_templates(templates):
    with open(TEMPLATE_FILE, "w") as f:
        json.dump(templates, f, indent=4)
    st.cache_resource.clear()

def update_template_mapping(template_name, mappings):
    templates = load_templates()
    for t in templates:
        if t["template_name"] == template_name:
            if t.get("mappings") != mappings:
                t["mappings"] = mappings
                save_templates(templates)
            return

def get_template_mapping(template_obj):
    return template_obj.get("mappings", {})

def map_column(label, columns, saved, key_prefix):
    mapping_type = saved.get("type", "column") if isinstance(saved, dict) else "column"
    value = saved.get("value", "") if isinstance(saved, dict) else saved

    mapping_type = st.radio(
        f"{label} - Map type", ["Column", "Custom"], 
        index=(0 if mapping_type=="column" else 1),
        key=f"{key_prefix}_type",
        horizontal=True  # <-- Make radio horizontal
    )
    result = {}
    if mapping_type == "Column":
        idx = columns.get_loc(value) if value in columns else 0
        selected_col = st.selectbox(
            "", options=columns, index=idx, key=f"{key_prefix}_col",
            label_visibility="collapsed"
        )
        result = {"type": "column", "value": selected_col}
    else:
        custom_val = st.text_input(
            "", value=value, key=f"{key_prefix}_custom",
            label_visibility="collapsed"
        )
        result = {"type": "custom", "value": custom_val}
    return result

# ---------- App Layout ----------
st.set_page_config(page_title="📤 WhatsApp Sender", layout="centered")
st.title("📤 WhatsApp Message Sender")

tab1, tab2 = st.tabs(["🧹 Manage Templates", "📤 Send WhatsApp Messages"])

# ========================================
# TAB 1: Manage Templates
# ========================================
with tab1:
    templates = get_templates_cached()

    st.subheader("➕ Add New Template")
    with st.form("template_form"):
        t_name = st.text_input("Template Name")
        t_id = st.text_input("Template ID (as per API)")
        t_msg = st.text_area("Message with placeholders like {{name}}, {{code}}", height=150)
        if st.form_submit_button("✅ Save Template"):
            if t_name and t_msg and t_id:
                templates.append({
                    "template_name": t_name,
                    "template_id": t_id,
                    "message": t_msg,
                    "mappings": {}
                })
                save_templates(templates)
                st.success("✅ Template added. Refresh to view.")

    st.subheader("📋 All Templates")
    if templates:
        for idx, t in enumerate(templates):
            with st.expander(f"🧾 {t['template_name']} (ID: {t['template_id']})"):
                new_name = st.text_input(f"Template Name ({idx})", value=t['template_name'], key=f"name_{idx}")
                new_id = st.text_input(f"Template ID ({idx})", value=t['template_id'], key=f"id_{idx}")
                new_msg = st.text_area(f"Template Message ({idx})", value=t['message'], height=120, key=f"msg_{idx}")

                col1, col2 = st.columns(2)
                with col1:
                    if st.button("💾 Update", key=f"update_{idx}"):
                        templates[idx] = {
                            "template_name": new_name,
                            "template_id": new_id,
                            "message": new_msg,
                            "mappings": t.get("mappings", {})
                        }
                        save_templates(templates)
                        st.success("✅ Template updated.")

                with col2:
                    if st.button("🗑️ Delete", key=f"delete_{idx}"):
                        templates.pop(idx)
                        save_templates(templates)
                        st.warning("🗑️ Template deleted.")
                        st.experimental_rerun()
    else:
        st.info("No templates found.")

# ========================================
# TAB 2: Send WhatsApp Messages
# ========================================
with tab2:
    uploaded_file = st.file_uploader("Upload Excel File", type=["xlsx"])

    if uploaded_file:
        df = pd.read_excel(uploaded_file)
        st.write("📄 File Preview", df.head())

        templates = get_templates_cached()
        if not templates:
            st.warning("No templates found. Please add a template in 'Manage Templates' tab.")
            st.stop()

        selected_template = st.selectbox("📋 Select Template to Use", [t["template_name"] for t in templates])
        template_obj = next((t for t in templates if t["template_name"] == selected_template), None)

        if template_obj:
            required_fields = re.findall(r"\{\{(.*?)\}\}", template_obj["message"])
            st.info(f"🔑 This template requires: {', '.join(required_fields)}")

            saved_mapping = get_template_mapping(template_obj)
            column_mapping = {}

            st.subheader("📌 Map Excel Columns or Manual Value")
            column_mapping["mobile_no"] = map_column(
                "📱 Mobile Number Column", df.columns, saved_mapping.get("mobile_no", {}), "mobile_map"
            )

            for field in required_fields:
                column_mapping[field] = map_column(
                    f"Map {{ {field} }} to column/text:", df.columns, saved_mapping.get(field, {}), f"map_{field}"
                )

            st.subheader("🖼️ Image Upload")
            image_col_map = st.selectbox(
                "Optional: Select image column", [""] + df.columns.tolist(),
                index=([""] + df.columns.tolist()).index(saved_mapping.get("image_column", "")) 
                if saved_mapping.get("image_column", "") in df.columns else 0
            )
            uploaded_image = st.file_uploader("Upload default image (optional)", type=["jpg", "jpeg", "png"])
            image_url = ""
            if uploaded_image:
                with st.spinner("Uploading image to imgBB..."):
                    imgbb_api_key = "085bfec8c1691df489038be3dce95cfc"
                    response = requests.post(
                        f"https://api.imgbb.com/1/upload?key={imgbb_api_key}",
                        files={"image": uploaded_image}
                    )
                    if response.status_code == 200:
                        data = response.json()['data']
                        image_url = data.get('display_url') or data.get('url') or data.get('image', {}).get('url')
                        st.success("✅ Image uploaded!")
                        st.image(image_url, caption="Default Image", use_container_width=True)
                        st.markdown(f"**Direct Image Link:**\n[{image_url}]({image_url})")
                    else:
                        st.error("❌ Failed to upload image.")

            st.subheader("🔍 Message Preview")
            try:
                preview_message = template_obj["message"]
                row = df.iloc[0]
                for field, mapinfo in column_mapping.items():
                    if field == "mobile_no":
                        continue
                    if mapinfo["type"] == "column":
                        val = str(row.get(mapinfo["value"], "")).strip()
                    else:
                        val = mapinfo["value"]
                    preview_message = preview_message.replace(f"{{{{{field}}}}}", val)
                st.markdown(f"**To:** {row.get(column_mapping['mobile_no']['value'], '')}")
                st.code(preview_message)
            except Exception as e:
                st.error(f"Preview error: {e}")

            if st.button("🚀 Send WhatsApp Messages"):
                mapping_to_save = column_mapping.copy()
                if image_col_map:
                    mapping_to_save["image_column"] = image_col_map
                update_template_mapping(template_obj["template_name"], mapping_to_save)

                url = "https://cloud.yellow.ai/api/engagements/notifications/v2/push?bot=x1683181251134"
                api_key = "zraDR0fGJ3AKLGM0HYauI5j-UXRaDZAzwGUVTvoZ"
                headers = {"x-api-key": api_key, "Content-Type": "application/json"}

                status = [""] * len(df)
                response_texts = [""] * len(df)

                def send_message(i, row):
                    try:
                        # PRIORITY LOGIC: Uploaded image > per-row image column > no image
                        final_img = ""
                        if image_url:
                            final_img = image_url
                        elif image_col_map:
                            row_img_url = str(row.get(image_col_map, "")).strip()
                            if row_img_url.startswith("http"):
                                final_img = row_img_url

                        # Build params using flexible mapping
                        params = {}
                        for field, mapinfo in column_mapping.items():
                            if field == "mobile_no":
                                continue
                            if mapinfo["type"] == "column":
                                val = str(row.get(mapinfo["value"], "")).strip()
                            else:
                                val = mapinfo["value"]
                            params[field] = val

                        payload = {
                            "userDetails": {
                                "number": str(row.get(column_mapping["mobile_no"]["value"], ""))
                            },
                            "notification": {
                                "type": "whatsapp",
                                "sender": "919311239211",
                                "templateId": template_obj["template_id"],
                                "params": params
                            }
                        }

                        if final_img:
                            params["media"] = {"mediaLink": final_img}

                        response = requests.post(url, json=payload, headers=headers)
                        resp_json = response.json()
                        if response.status_code in [200, 202]:
                            return (i, "Sent", resp_json.get("message", "Success"))
                        else:
                            return (i, "Failed", resp_json.get("message", f"HTTP {response.status_code}"))
                    except Exception as e:
                        return (i, "Error", str(e))

                st.subheader("📊 Sending Messages...")
                progress_bar = st.progress(0)
                completed = 0
                total = len(df)

                with ThreadPoolExecutor(max_workers=10) as executor:
                    futures = [executor.submit(send_message, i, row) for i, row in df.iterrows()]
                    for future in as_completed(futures):
                        i, stat, msg = future.result()
                        status[i] = stat
                        response_texts[i] = msg
                        completed += 1
                        progress_bar.progress(completed / total)

                df["Status"] = status
                df["API Response"] = response_texts
                st.success("✅ All messages processed.")
                st.dataframe(df)

                timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
                output_file = f"Message_Report_{timestamp}.xlsx"
                df.to_excel(output_file, index=False)

                wb = load_workbook(output_file)
                ws = wb.active
                red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                for row_idx in range(2, ws.max_row + 1):
                    if str(ws[f"D{row_idx}"].value).lower() in ["failed", "error"]:
                        for col_idx in range(1, ws.max_column + 1):
                            ws.cell(row=row_idx, column=col_idx).fill = red_fill
                wb.save(output_file)

                subprocess.Popen(f'explorer "{os.path.abspath(os.path.dirname(output_file))}"')

                with open(output_file, "rb") as f:
                    st.download_button("📥 Download Final Report", f, file_name=os.path.basename(output_file))
